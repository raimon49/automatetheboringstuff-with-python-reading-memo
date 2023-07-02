#!/usr/bin/env python
# -*- coding: utf-8 -*-
# vim:fenc=utf-8 ff=unix ft=python ts=4 sw=4 sts=4 si et

def main():
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string

    wb = openpyxl.load_workbook('produceSales.xlsx')
    print(type(wb))
    print(wb.sheetnames)

    sheet = wb['Sheet']
    print(type(sheet))

    a1 = sheet['A1']
    print(a1.value)
    b1 = sheet['B1']
    print(b1.value)

    for i in range(1, 8, 2):
        print(i, sheet.cell(row=i, column=2).value)

    print(sheet.max_row)
    print(sheet.max_column)

    print(get_column_letter(1))   # A
    print(get_column_letter(2))   # B
    print(get_column_letter(27))  # AA
    print(get_column_letter(900)) # AHP

    print(get_column_letter(sheet.max_column)) # D

    print(column_index_from_string('A'))  # 1
    print(column_index_from_string('AA')) # 27
    print(column_index_from_string('AA')) # 27

    for row_of_cell_objects in sheet['A1':'C3']: # 矩形領域をスライスで取得
        for cell_obj in row_of_cell_objects:     # 行に含まれるセルを取得
            print(cell_obj.coordinate, cell_obj.value)

        print('--- END OF ROw ---')

if __name__ == '__main__':
    main()

