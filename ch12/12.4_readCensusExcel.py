#!/usr/bin/env python
# -*- coding: utf-8 -*-
# vim:fenc=utf-8 ff=unix ft=python ts=4 sw=4 sts=4 si et

def main():
    import openpyxl, pprint

    print('ワークブックを開いています...')
    wb = openpyxl.load_workbook('censuspopdata.xlsx')
    print(wb.sheetnames)
    sheet = wb['Population by Census Tract']
    country_data = {}

    print('行を読み込んでいます...')
    for row in range(2, sheet.max_row + 1):
        # スプレッドシートの1行に、ひとつの人口調査標準地域のデータがある
        state = sheet['B' + str(row)].value
        country = sheet['C' + str(row)].value
        pop = sheet['D' + str(row)].value

        # この州のキーが確実に存在するようにする
        country_data.setdefault(state, {})
        # この州のこの郡のキーが確実に存在するようにする
        country_data[state].setdefault(country, {'tracts': 0, 'pop': 0})
        # 各行が人口調査標準地域を表すので、数を1つ増やす
        country_data[state][country]['tracts'] += 1
        # この人口調査標準地域の人口だけ郡の人口を増やす
        country_data[state][country]['pop'] += int(pop)

    print('結果を描き込み中...')
    with open('census2010.py', 'w') as result_file:
        result_file.write('all_data = ' + pprint.pformat(country_data))

    print('完了')
    # Pythonインタラクティブシェルで
    # >>> import census2010
    # >>> census2010.all_data['AK']['Anchroge']
    # >>> census2010.all_data['AK']['Aleutians West']
    # でデータを確認可能

if __name__ == '__main__':
    main()

