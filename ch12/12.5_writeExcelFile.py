#!/usr/bin/env python
# -*- coding: utf-8 -*-
# vim:fenc=utf-8 ff=unix ft=python ts=4 sw=4 sts=4 si et

def main():
    import openpyxl

    wb = openpyxl.Workbook()
    print(wb.sheetnames)

    sheet = wb.active
    # シートの名前を変更
    sheet.title = 'Spam Bacon Eggs Sheet'
    print(wb.sheetnames)

    # デフォルトの'Sheet'という名称で末尾にシート追加
    wb.create_sheet()
    print(wb.sheetnames)

    # 追加する位置とシート名を指定してシート追加（先頭）
    wb.create_sheet(index=0, title='First Sheet')
    print(wb.sheetnames)

    # 追加する位置とシート名を指定してシート追加（3番目）
    wb.create_sheet(index=2, title='Middle Sheet')
    print(wb.sheetnames)

    # シート名を指定してワークブックから削除
    del wb['Middle Sheet']
    del wb['Spam Bacon Eggs Sheet']
    print(wb.sheetnames)

    # セルに値を書き込む
    wb['First Sheet']['A1'] = 'Hello, World!'
    print(wb['First Sheet']['A1'].value)

if __name__ == '__main__':
    main()

