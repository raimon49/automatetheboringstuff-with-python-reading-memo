#!/usr/bin/env python
# -*- coding: utf-8 -*-
# vim:fenc=utf-8 ff=unix ft=python ts=4 sw=4 sts=4 si et

def main():
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.load_workbook('produceSales.xlsx', data_only=False) # 数式を計算結果として開くときはdata_only=Trueを指定する
    sheet = wb['Sheet']

    # セルA1のフォントスタイルを変更する
    italic24_font = Font(size=24, italic=True)
    sheet['A1'].font = italic24_font
    sheet['A1'] = 'Hello, italic world!!'

    # セルA2のフォントスタイルを変更する
    bold24_font = Font(name='Times New Roman', size=24, bold=True)
    sheet['A2'].font = bold24_font
    sheet['A2'] = 'Bold Times New Roman'

    # 1行目の高さを設定
    sheet.row_dimensions[1].height = 70
    # 列Bの幅を設定
    sheet.column_dimensions['B'].width = 20

    # 数式の適用
    sheet['B9'] = '=SUM(B1:B8)'

    # セルC10:D12の範囲を結合
    sheet.merge_cells('C10:D12')

    # セルの結合を解除
    sheet.unmerge_cells('C10:D12')

    # ウィンドウ枠の固定（1行目を見出し化）
    sheet.freeze_panes = 'A2'

    # ウィンドウ枠の固定を解除
    sheet.freeze_panes = None

    # 農産物の種類と、更新する価格
    PRICE_UPDATES = {'Garlic': 3.07,
                     'Celery': 1.19,
                     'Lemon': 1.27}

    # 行をループして価格を更新する
    for row_num in range(2, sheet.max_row + 1): # 先頭行は表見出しのためスキップ
        produce_name = sheet.cell(row = row_num, column=1).value
        if produce_name in PRICE_UPDATES:
            sheet.cell(row=row_num, column=2).value = PRICE_UPDATES[produce_name]

    # 間違った内容で更新されないよう、元のファイルを上書きせず別名で保存
    wb.save('updatedProduceSales.xlsx')

if __name__ == '__main__':
    main()

