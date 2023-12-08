#!/usr/bin/env python
# -*- coding: utf-8 -*-
# vim:fenc=utf-8 ff=unix ft=python ts=4 sw=4 sts=4 si et

def main():
    import openpyxl, smtplib, sys

    wb = openpyxl.load_workbook('duesRecords.xlsx')
    sheet = wb['Sheet1']

    last_col = sheet.max_column
    latest_month = sheet.cell(row=1, column=last_col).value # 最新月を取得

    # 会員の支払い状況を調べる
    unpaied_members = {}
    for r in range(2, sheet.max_row + 1): # 2行目以降をループ
        payment = sheet.cell(row=r, column=last_col).value # 'paid' or None
        if payment != 'paid':
            # 支払い済みでない会員の情報を取得
            name = sheet.cell(row=r, column=1).value
            email = sheet.cell(row=r, column=2).value
            unpaied_members[name] = email # 名前をキーにメールアドレスを登録
    # メールアカウントにログインする
    smtp_obj = smtplib.SMTP('smtp.gmail.com', 587)
    smtp_obj.ehlo()
    smtp_obj.starttls()
    smtp_obj.login('my_email_address@gmail.com', sys.argv[1]) # パスワードは引数で受け取る
    # TODO: リマインダーメールを送信する
    for name, email in unpaied_members.items():
        body = """Subject:  {} dues unpaid.
Dear {},
Records show that you have not paid dues for {}. Please make this payment as soon as possible. Thank you!
""".format(latest_month, name, latest_month)

        print('メール送信中 {}...'.format(email))
        sendmail_status = smtp_obj.sendmail('my_email_address@gmail.com', email, body)


if __name__ == '__main__':
    main()
